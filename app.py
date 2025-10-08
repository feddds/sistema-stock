from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from models import db, Insumo, Compra, Consumo, Usuario
from models import CentroConsumo, Trabajador
from datetime import datetime
from functools import wraps
#from flask_migrate import Migrate
import os

from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import openpyxl  # ← AGREGAR ESTA IMPORTACIÓN
import openpyxl.utils  # ← AGREGAR ESTA IMPORTACIÓN
import io
import os  # ← MOVER AQUÍ

app = Flask(__name__)
# POR ESTA (para producción):
if os.environ.get('FLASK_ENV') == 'production':
    # En producción (PythonAnywhere) usa path absoluto
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.db')
else:
    # En desarrollo local
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'

app.config['SECRET_KEY'] = 'clave_secreta_stock_app_2024'
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_TYPE'] = 'filesystem'

db.init_app(app)

#migrate = Migrate(app, db)

# === DECORADORES DE SEGURIDAD ===
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor inicia sesión para acceder a esta página.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Por favor inicia sesión para acceder a esta página.', 'warning')
                return redirect(url_for('login'))
            
            #user = Usuario.query.get(session['user_id'])
            user = db.session.get(Usuario, session['user_id'])

                        # 👇 AGREGAR ESTOS PRINTS PARA DEBUG
            print(f"🔍 DEBUG ROLE_REQUIRED:")
            print(f"   user_id en session: {session.get('user_id')}")
            print(f"   usuario encontrado: {user}")
            if user:
                print(f"   rol del usuario: {user.rol}")
                print(f"   roles permitidos: {roles}")
                print(f"   ¿tiene acceso?: {user.rol in roles}")
            
            if not user or user.rol not in roles:
                print(f"   ❌ ACCESO DENEGADO")
                flash('No tienes permisos para acceder a esta sección', 'danger')
                return redirect(url_for('index'))
            
            print(f"   ✅ ACCESO PERMITIDO")
            
            #if user.rol not in roles:
            if not user or user.rol not in roles:
                flash('No tienes permisos para acceder a esta función.', 'danger')
                return redirect(url_for('index'))
            

            return f(*args, **kwargs)
        return decorated_function
    return decorator

# === RUTAS DE AUTENTICACIÓN ===
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = Usuario.query.filter_by(username=username, activo=True).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['user_rol'] = user.rol
            session['user_nombre'] = user.nombre
            
            flash(f'¡Bienvenido {user.nombre}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión correctamente', 'info')
    return redirect(url_for('login'))

@app.route('/')
@login_required  # ← AGREGAR ESTA LÍNEA
def index():
    return render_template('index.html')

@app.route('/buscar_insumos')
@login_required
def buscar_insumos():
    """Buscar insumos para autocomplete"""
    query = request.args.get('q', '').lower().strip()
    
    if not query or len(query) < 2:
        return jsonify([])
    
    # Buscar en denominación, tipo y modelo
    insumos = Insumo.query.filter(
        db.or_(
            Insumo.denominacion.ilike(f'%{query}%'),
            Insumo.tipo.ilike(f'%{query}%'),
            Insumo.modelo.ilike(f'%{query}%')
        )
    ).limit(20).all()
    
    resultados = []
    for insumo in insumos:
        resultados.append({
            'id': insumo.id,
            'text': f"{insumo.denominacion} - {insumo.tipo} - {insumo.modelo}",
            'denominacion': insumo.denominacion,
            'tipo': insumo.tipo,
            'modelo': insumo.modelo,
            'stock_actual': insumo.stock_actual,
            'cantidad_por_caja': insumo.cantidad_por_caja
        })
    
    return jsonify(resultados)

@app.route('/get_trabajadores_por_centro/<int:centro_id>')
@login_required
def get_trabajadores_por_centro(centro_id):
    """Obtener trabajadores activos por centro"""
    trabajadores = Trabajador.query.filter_by(
        centro_consumo_id=centro_id, 
        activo=True
    ).order_by(Trabajador.nombre).all()
    
    resultados = []
    for trabajador in trabajadores:
        resultados.append({
            'id': trabajador.id,
            'nombre': trabajador.nombre,
            'codigo': trabajador.codigo,
            'centro_nombre': trabajador.centro_consumo.nombre
        })
    
    return jsonify(resultados)

@app.route('/crear-insumo', methods=['GET', 'POST'])
@login_required  # ← AGREGAR ESTA LÍNEA
@role_required(['admin'])  # ← AGREGAR ESTA LÍNEA - Solo admin puede crear insumos
def crear_insumo():
    if request.method == 'POST':
        try:
            nuevo_insumo = Insumo(
                denominacion=request.form['denominacion'],
                tipo=request.form['tipo'],
                modelo=request.form['modelo'],
                cantidad_por_caja=int(request.form['cantidad_por_caja']),
                precio_caja=float(request.form['precio_caja']),
                precio_unitario=float(request.form['precio_unitario']),
                codigo_barras=request.form['codigo_barras'] or None
            )
            
            db.session.add(nuevo_insumo)
            db.session.commit()
            flash('Insumo creado exitosamente!', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear insumo: {str(e)}', 'danger')
    
    return render_template('crear_insumo.html')

@app.route('/registrar-compra', methods=['GET', 'POST'])
@login_required  # ← AGREGAR ESTA LÍNEA
@role_required(['compras', 'admin'])  # ← AGREGAR ESTA LÍNEA - Compras y admin
def registrar_compra():
    insumos = Insumo.query.all()
    
    if request.method == 'POST':
        try:
            fecha_vencimiento = request.form['fecha_vencimiento']
            fecha_vencimiento = datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date() if fecha_vencimiento else None
            
            nueva_compra = Compra(
                insumo_id=int(request.form['insumo_id']),
                cantidad_cajas=float(request.form['cantidad_cajas']),
                precio_caja_compra=float(request.form['precio_caja_compra']),
                proveedor=request.form['proveedor'],
                lote=request.form['lote'],
                fecha_vencimiento=fecha_vencimiento
            )
            
            db.session.add(nueva_compra)
            db.session.commit()
            flash('Compra registrada exitosamente!', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar compra: {str(e)}', 'danger')
    
    return render_template('registrar_compra.html', insumos=insumos)

@app.route('/registrar-consumo', methods=['GET', 'POST'])
@login_required  # ← AGREGAR ESTA LÍNEA
@role_required(['stock', 'admin'])  # ← AGREGAR ESTA LÍNEA - Stock y admin
def registrar_consumo():
    insumos = Insumo.query.all()
    centros = CentroConsumo.query.filter_by(activo=True).all()  # ← NUEVO
    
    if request.method == 'POST':
        try:
            # Validar que el insumo existe
            insumo = Insumo.query.get(int(request.form['insumo_id']))
            if not insumo:
                flash('❌ El insumo seleccionado no existe', 'error')
                return redirect(url_for('registrar_consumo'))
            
            # Validar stock suficiente
            cantidad = float(request.form['cantidad_unidades'])
            if cantidad <= 0:
                flash('❌ La cantidad debe ser mayor a cero', 'error')
                return redirect(url_for('registrar_consumo'))
                
            if insumo.stock_actual < cantidad:
                flash(f'❌ Stock insuficiente. Disponible: {insumo.stock_actual} unidades', 'error')
                return redirect(url_for('registrar_consumo'))
            
            # Validar centro y trabajador
            centro = CentroConsumo.query.get(int(request.form['centro_consumo_id']))
            trabajador = Trabajador.query.get(int(request.form['trabajador_id']))
            
            if not centro:
                flash('❌ El centro de consumo seleccionado no existe', 'error')
                return redirect(url_for('registrar_consumo'))
                
            if not trabajador:
                flash('❌ El trabajador seleccionado no existe', 'error')
                return redirect(url_for('registrar_consumo'))
            
            # Verificar que el trabajador pertenezca al centro
            if trabajador.centro_consumo_id != centro.id:
                flash('❌ El trabajador no pertenece al centro seleccionado', 'error')
                return redirect(url_for('registrar_consumo'))
            
            # Crear el consumo con los nuevos campos
            nuevo_consumo = Consumo(
                insumo_id=int(request.form['insumo_id']),
                cantidad_unidades=cantidad,
                proyecto=request.form['proyecto'],
                observaciones=request.form['observaciones'],
                centro_consumo_id=centro.id,      # ← NUEVO
                trabajador_id=trabajador.id       # ← NUEVO
                # ELIMINADO: responsable=request.form['responsable']
            )
            
            db.session.add(nuevo_consumo)
            db.session.commit()
            
            flash(f'✅ Consumo registrado exitosamente! Stock actual: {insumo.stock_actual - cantidad:.2f} unidades', 'success')
            return redirect(url_for('registrar_consumo'))
            
        except ValueError:
            flash('❌ Error en los datos ingresados. Verifique los valores.', 'error')
            return redirect(url_for('registrar_consumo'))
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error al registrar el consumo: {str(e)}', 'error')
            return redirect(url_for('registrar_consumo'))
    
    return render_template('registrar_consumo.html', insumos=insumos, centros=centros)

@app.route('/reporte-stock')
@login_required  # ← AGREGAR ESTA LÍNEA
@role_required(['basico', 'stock', 'compras', 'admin'])  # ← AGREGAR ESTA LÍNEA - Todos pueden ver
def reporte_stock():
    insumos = Insumo.query.all()
    reporte = []
    
    for insumo in insumos:
        total_unidades_compradas = sum(
            compra.cantidad_cajas * insumo.cantidad_por_caja 
            for compra in insumo.compras
        )
        
        total_unidades_consumidas = sum(
            consumo.cantidad_unidades for consumo in insumo.consumos
        )
        
        stock_actual_unidades = total_unidades_compradas - total_unidades_consumidas
        cajas_completas = stock_actual_unidades // insumo.cantidad_por_caja
        unidades_sueltas = stock_actual_unidades % insumo.cantidad_por_caja
        valor_stock = stock_actual_unidades * insumo.precio_unitario
        
        reporte.append({
            'insumo': insumo,
            'total_unidades_compradas': total_unidades_compradas,
            'total_unidades_consumidas': total_unidades_consumidas,
            'stock_actual_unidades': stock_actual_unidades,
            'cajas_completas': int(cajas_completas),
            'unidades_sueltas': unidades_sueltas,
            'valor_stock': valor_stock
        })
    
    return render_template('reporte_stock.html', reporte=reporte)


@app.route('/exportar_consumos_excel')
def exportar_consumos_excel():
    """Exportar consumos a Excel"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Obtener todos los consumos
        consumos = Consumo.query.order_by(Consumo.fecha_consumo.desc()).all()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Consumos"
        
        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'Fecha', 'Hora', 'Insumo', 'Tipo', 'Modelo', 
            'Cantidad Unidades', 'Cajas Equivalentes', 'Centro Consumo', 
            'Trabajador', 'Código Trabajador', 'Proyecto', 'Observaciones', 
            'Costo', 'Costo Total'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        for row, consumo in enumerate(consumos, 2):
            ws.cell(row=row, column=1, value=consumo.fecha_consumo.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=consumo.fecha_consumo.strftime('%H:%M'))
            ws.cell(row=row, column=3, value=consumo.insumo.denominacion)
            ws.cell(row=row, column=4, value=consumo.insumo.tipo)
            ws.cell(row=row, column=5, value=consumo.insumo.modelo)
            ws.cell(row=row, column=6, value=float(consumo.cantidad_unidades))
            ws.cell(row=row, column=7, value=float(consumo.cantidad_cajas_equivalentes))
            ws.cell(row=row, column=8, value=consumo.centro_consumo.nombre)
            ws.cell(row=row, column=9, value=consumo.trabajador.nombre)
            ws.cell(row=row, column=10, value=consumo.trabajador.codigo)
            ws.cell(row=row, column=11, value=consumo.proyecto or '')
            ws.cell(row=row, column=12, value=consumo.observaciones or '')
            ws.cell(row=row, column=13, value=float(consumo.costo_consumo))
            ws.cell(row=row, column=14, value=float(consumo.costo_consumo))
        
        # Ajustar anchos de columnas
        column_widths = [12, 8, 25, 15, 12, 12, 12, 15, 20, 12, 15, 25, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Totales
        if consumos:
            total_row = len(consumos) + 3
            ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{len(consumos)+1})")
            ws.cell(row=total_row, column=14, value=f"=SUM(N2:N{len(consumos)+1})")
            ws.cell(row=total_row, column=5, value="TOTALES:").font = Font(bold=True)
        
        # Guardar en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Crear respuesta
        fecha_exportacion = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"consumos_{fecha_exportacion}.xlsx"
        
        return Response(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        flash(f'❌ Error al exportar a Excel: {str(e)}', 'error')
        return redirect(url_for('listado_consumos'))

@app.route('/exportar_stock_excel')
def exportar_stock_excel():
    """Exportar stock actual a Excel"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        insumos = Insumo.query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Actual"
        
        # Encabezados
        headers = [
            'Insumo', 'Tipo', 'Modelo', 'Stock Actual', 'Stock Mínimo', 
            'Cajas Completas', 'Unidades Sueltas', 'Precio Unitario', 
            'Valor Stock', 'Estado Alerta', 'Porcentaje Stock'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        for row, insumo in enumerate(insumos, 2):
            estado = "🔴 CRÍTICO" if insumo.necesita_reposicion else "✅ OK" if insumo.stock_minimo > 0 else "⚪ SIN ALERTA"
            
            ws.cell(row=row, column=1, value=insumo.denominacion)
            ws.cell(row=row, column=2, value=insumo.tipo)
            ws.cell(row=row, column=3, value=insumo.modelo)
            ws.cell(row=row, column=4, value=float(insumo.stock_actual))
            ws.cell(row=row, column=5, value=float(insumo.stock_minimo))
            ws.cell(row=row, column=6, value=int(insumo.stock_en_cajas))
            ws.cell(row=row, column=7, value=int(insumo.unidades_sueltas))
            ws.cell(row=row, column=8, value=float(insumo.precio_unitario))
            ws.cell(row=row, column=9, value=float(insumo.valor_stock_actual))
            ws.cell(row=row, column=10, value=estado)
            ws.cell(row=row, column=11, value=float(insumo.porcentaje_stock))
        
        # Ajustar columnas
        column_widths = [25, 15, 12, 12, 12, 12, 12, 12, 12, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Guardar
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        fecha_exportacion = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"stock_actual_{fecha_exportacion}.xlsx"
        
        return Response(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        flash(f'❌ Error al exportar stock a Excel: {str(e)}', 'error')
        return redirect(url_for('reporte_stock'))

@app.route('/gestion_insumos')
def gestion_insumos():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Solo admin puede gestionar insumos
    if session.get('user_rol') != 'admin':
        flash('No tienes permisos para acceder a esta sección', 'danger')
        return redirect(url_for('index'))
    
    insumos = Insumo.query.order_by(Insumo.denominacion, Insumo.tipo).all()
    return render_template('gestion_insumos.html', insumos=insumos)


@app.route('/get_consumo/<int:consumo_id>')
def get_consumo(consumo_id):
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'error': 'No autorizado'}), 403
    
    consumo = Consumo.query.get_or_404(consumo_id)
    return jsonify({
        'id': consumo.id,
        'insumo_id': consumo.insumo_id,
        'cantidad_unidades': float(consumo.cantidad_unidades),
        'proyecto': consumo.proyecto,
        'observaciones': consumo.observaciones,
        'centro_consumo_id': consumo.centro_consumo_id,
        'trabajador_id': consumo.trabajador_id,
        'fecha_consumo': consumo.fecha_consumo.isoformat(),
        'costo_consumo': float(consumo.costo_consumo)
    })


@app.route('/get_insumo/<int:insumo_id>')
@login_required
@role_required(['admin'])
def get_insumo(insumo_id):
    """Obtener datos de un insumo para editar"""
    insumo = Insumo.query.get_or_404(insumo_id)
    return jsonify({
        'id': insumo.id,
        'denominacion': insumo.denominacion,
        'tipo': insumo.tipo,
        'modelo': insumo.modelo,
        'cantidad_por_caja': float(insumo.cantidad_por_caja),
        'precio_caja': float(insumo.precio_caja),
        'precio_unitario': float(insumo.precio_unitario),
        'codigo_barras': insumo.codigo_barras or '',
        'stock_minimo': float(insumo.stock_minimo) if insumo.stock_minimo else 0
    })

@app.route('/editar_insumo', methods=['POST'])
@login_required
@role_required(['admin'])
def editar_insumo():
    """Editar insumo existente"""
    try:
        insumo_id = request.form.get('id')
        insumo = Insumo.query.get_or_404(insumo_id)
        
        # Actualizar datos
        insumo.denominacion = request.form.get('denominacion')
        insumo.tipo = request.form.get('tipo')
        insumo.modelo = request.form.get('modelo')
        insumo.cantidad_por_caja = float(request.form.get('cantidad_por_caja'))
        insumo.precio_caja = float(request.form.get('precio_caja'))
        insumo.precio_unitario = float(request.form.get('precio_unitario'))
        insumo.codigo_barras = request.form.get('codigo_barras') or None
        insumo.stock_minimo = float(request.form.get('stock_minimo')) if request.form.get('stock_minimo') else None
        
        db.session.commit()
        flash('✅ Insumo actualizado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('❌ Error al actualizar el insumo: ' + str(e), 'error')
    
    return redirect(url_for('gestion_insumos'))

@app.route('/actualizar_stock_minimo/<int:insumo_id>', methods=['POST'])
@login_required
@role_required(['admin'])
def actualizar_stock_minimo(insumo_id):
    """Actualizar solo el stock mínimo de un insumo"""
    try:
        insumo = Insumo.query.get_or_404(insumo_id)
        stock_minimo = request.form.get('stock_minimo')
        
        if stock_minimo and float(stock_minimo) >= 0:
            insumo.stock_minimo = float(stock_minimo)
        else:
            insumo.stock_minimo = None
        
        db.session.commit()
        return jsonify({'success': True, 'nuevo_stock_minimo': insumo.stock_minimo})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/editar_consumo', methods=['POST'])
def editar_consumo():
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        flash('No tenés permisos para esta acción', 'error')
        return redirect(url_for('listado_consumos'))
    
    try:
        consumo_id = request.form.get('id')
        consumo = Consumo.query.get_or_404(consumo_id)
        
        # Actualizar datos
        consumo.insumo_id = int(request.form.get('insumo_id'))
        consumo.cantidad_unidades = float(request.form.get('cantidad_unidades'))
        consumo.proyecto = request.form.get('proyecto')
        consumo.observaciones = request.form.get('observaciones')
        consumo.centro_consumo_id = int(request.form.get('centro_consumo_id'))
        consumo.trabajador_id = int(request.form.get('trabajador_id'))
        
        db.session.commit()
        flash('✅ Consumo actualizado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('❌ Error al actualizar el consumo: ' + str(e), 'error')
    
    return redirect(url_for('listado_consumos'))

@app.route('/eliminar_insumo/<int:insumo_id>', methods=['POST'])
def eliminar_insumo(insumo_id):
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        insumo = Insumo.query.get_or_404(insumo_id)
        
        # Verificar si hay stock antes de eliminar
        if insumo.stock_actual > 0:
            return jsonify({
                'success': False, 
                'error': 'No se puede eliminar un insumo con stock existente'
            })
        
        db.session.delete(insumo)
        db.session.commit()
        return jsonify({'success': True})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/listado_consumos')
@login_required
@role_required(['stock', 'admin'])
def listado_consumos():
    # Obtener todos los consumos ordenados por fecha más reciente primero
    consumos = Consumo.query.order_by(Consumo.fecha_consumo.desc()).all()
    
    # Datos para filtros
    centros_unicos = list(set([c.centro_consumo.nombre for c in consumos]))
    trabajadores_unicos = list(set([c.trabajador.nombre for c in consumos]))
    
    # Datos para edición (solo si es admin)
    insumos = Insumo.query.all() if session.get('user_rol') == 'admin' else []
    centros = CentroConsumo.query.filter_by(activo=True).all() if session.get('user_rol') == 'admin' else []
    
    return render_template('listado_consumos.html', 
                         consumos=consumos,
                         centros_unicos=centros_unicos,
                         trabajadores_unicos=trabajadores_unicos,
                         insumos=insumos,
                         centros=centros)


@app.route('/alertas_stock')
def alertas_stock():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Obtener todos los insumos
    insumos = Insumo.query.all()
    
    # Filtrar por diferentes estados
    alertas_criticas = [i for i in insumos if i.stock_minimo > 0 and i.necesita_reposicion]
    insumos_con_alerta = [i for i in insumos if i.stock_minimo > 0]
    insumos_ok = [i for i in insumos_con_alerta if not i.necesita_reposicion]
    insumos_sin_alerta = [i for i in insumos if i.stock_minimo == 0]
    
    return render_template('alertas_stock.html',
                         alertas_criticas=alertas_criticas,
                         insumos_con_alerta=insumos_con_alerta,
                         insumos_ok=insumos_ok,
                         insumos_sin_alerta=insumos_sin_alerta)


@app.route('/listado-compras')
@login_required  # ← AGREGAR ESTA LÍNEA
@role_required(['compras', 'admin'])  # ← AGREGAR ESTA LÍNEA - Compras y admin
def listado_compras():
    # Obtener todas las compras ordenadas por fecha más reciente primero
    compras = Compra.query.order_by(Compra.fecha_compra.desc()).all()
    return render_template('listado_compras.html', compras=compras)


# Añadir estas rutas a tu app.py o routes.py

@app.route('/gestion_centros')
def gestion_centros():
    """Vista principal de gestión de centros de consumo"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        flash('No tenés permisos para acceder a esta sección', 'error')
        return redirect(url_for('index'))
    
    centros = CentroConsumo.query.all()
    return render_template('gestion_centros.html', centros=centros)

@app.route('/crear_centro', methods=['POST'])
def crear_centro():
    """Crear nuevo centro de consumo"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        nombre = request.form.get('nombre')
        descripcion = request.form.get('descripcion')
        activo = request.form.get('activo') == 'on'
        
        # Validar que el nombre no exista
        centro_existente = CentroConsumo.query.filter_by(nombre=nombre).first()
        if centro_existente:
            flash('❌ Ya existe un centro con ese nombre', 'error')
            return redirect(url_for('gestion_centros'))
        
        nuevo_centro = CentroConsumo(
            nombre=nombre,
            descripcion=descripcion,
            activo=activo
        )
        
        db.session.add(nuevo_centro)
        db.session.commit()
        
        flash('✅ Centro creado exitosamente', 'success')
        return redirect(url_for('gestion_centros'))
        
    except Exception as e:
        db.session.rollback()
        flash('❌ Error al crear el centro: ' + str(e), 'error')
        return redirect(url_for('gestion_centros'))

@app.route('/get_centro/<int:centro_id>')
def get_centro(centro_id):
    """Obtener datos de un centro para editar"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'error': 'No autorizado'}), 403
    
    centro = CentroConsumo.query.get_or_404(centro_id)
    return jsonify({
        'id': centro.id,
        'nombre': centro.nombre,
        'descripcion': centro.descripcion,
        'activo': centro.activo,
        'trabajadores_count': len(centro.trabajadores),
        'consumos_count': len(centro.consumos),
        'created_at': centro.created_at.isoformat()
    })

@app.route('/editar_centro', methods=['POST'])
def editar_centro():
    """Editar centro de consumo existente"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        centro_id = request.form.get('id')
        nombre = request.form.get('nombre')
        descripcion = request.form.get('descripcion')
        activo = request.form.get('activo') == 'on'
        
        centro = CentroConsumo.query.get_or_404(centro_id)
        
        # Validar que el nombre no exista (excluyendo el actual)
        centro_existente = CentroConsumo.query.filter(
            CentroConsumo.nombre == nombre,
            CentroConsumo.id != centro_id
        ).first()
        if centro_existente:
            flash('❌ Ya existe otro centro con ese nombre', 'error')
            return redirect(url_for('gestion_centros'))
        
        centro.nombre = nombre
        centro.descripcion = descripcion
        centro.activo = activo
        
        db.session.commit()
        
        flash('✅ Centro actualizado exitosamente', 'success')
        return redirect(url_for('gestion_centros'))
        
    except Exception as e:
        db.session.rollback()
        flash('❌ Error al actualizar el centro: ' + str(e), 'error')
        return redirect(url_for('gestion_centros'))

@app.route('/desactivar_centro/<int:centro_id>', methods=['POST'])
def desactivar_centro(centro_id):
    """Desactivar un centro de consumo"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        centro = CentroConsumo.query.get_or_404(centro_id)
        centro.activo = False
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/activar_centro/<int:centro_id>', methods=['POST'])
def activar_centro(centro_id):
    """Activar un centro de consumo"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        centro = CentroConsumo.query.get_or_404(centro_id)
        centro.activo = True
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/gestion_trabajadores')
def gestion_trabajadores():
    """Vista principal de gestión de trabajadores"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        flash('No tenés permisos para acceder a esta sección', 'error')
        return redirect(url_for('index'))
    
    trabajadores = Trabajador.query.join(CentroConsumo).all()
    centros = CentroConsumo.query.all()
    return render_template('gestion_trabajadores.html', 
                         trabajadores=trabajadores, 
                         centros=centros)

@app.route('/crear_trabajador', methods=['POST'])
def crear_trabajador():
    """Crear nuevo trabajador"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        codigo = request.form.get('codigo')
        nombre = request.form.get('nombre')
        centro_consumo_id = request.form.get('centro_consumo_id')
        activo = request.form.get('activo') == 'on'
        
        # Validar que el código no exista
        trabajador_existente = Trabajador.query.filter_by(codigo=codigo).first()
        if trabajador_existente:
            flash('❌ Ya existe un trabajador con ese código', 'error')
            return redirect(url_for('gestion_trabajadores'))
        
        # Validar que el centro exista
        centro = CentroConsumo.query.get(centro_consumo_id)
        if not centro:
            flash('❌ El centro seleccionado no existe', 'error')
            return redirect(url_for('gestion_trabajadores'))
        
        nuevo_trabajador = Trabajador(
            codigo=codigo,
            nombre=nombre,
            centro_consumo_id=centro_consumo_id,
            activo=activo
        )
        
        db.session.add(nuevo_trabajador)
        db.session.commit()
        
        flash('✅ Trabajador creado exitosamente', 'success')
        return redirect(url_for('gestion_trabajadores'))
        
    except Exception as e:
        db.session.rollback()
        flash('❌ Error al crear el trabajador: ' + str(e), 'error')
        return redirect(url_for('gestion_trabajadores'))

@app.route('/get_trabajador/<int:trabajador_id>')
def get_trabajador(trabajador_id):
    """Obtener datos de un trabajador para editar"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'error': 'No autorizado'}), 403
    
    trabajador = Trabajador.query.get_or_404(trabajador_id)
    return jsonify({
        'id': trabajador.id,
        'codigo': trabajador.codigo,
        'nombre': trabajador.nombre,
        'centro_consumo_id': trabajador.centro_consumo_id,
        'centro_nombre': trabajador.centro_consumo.nombre,
        'activo': trabajador.activo,
        'consumos_count': len(trabajador.consumos),
        'created_at': trabajador.created_at.isoformat()
    })

@app.route('/editar_trabajador', methods=['POST'])
def editar_trabajador():
    """Editar trabajador existente"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        trabajador_id = request.form.get('id')
        codigo = request.form.get('codigo')
        nombre = request.form.get('nombre')
        centro_consumo_id = request.form.get('centro_consumo_id')
        activo = request.form.get('activo') == 'on'
        
        trabajador = Trabajador.query.get_or_404(trabajador_id)
        
        # Validar que el código no exista (excluyendo el actual)
        trabajador_existente = Trabajador.query.filter(
            Trabajador.codigo == codigo,
            Trabajador.id != trabajador_id
        ).first()
        if trabajador_existente:
            flash('❌ Ya existe otro trabajador con ese código', 'error')
            return redirect(url_for('gestion_trabajadores'))
        
        # Validar que el centro exista
        centro = CentroConsumo.query.get(centro_consumo_id)
        if not centro:
            flash('❌ El centro seleccionado no existe', 'error')
            return redirect(url_for('gestion_trabajadores'))
        
        trabajador.codigo = codigo
        trabajador.nombre = nombre
        trabajador.centro_consumo_id = centro_consumo_id
        trabajador.activo = activo
        
        db.session.commit()
        
        flash('✅ Trabajador actualizado exitosamente', 'success')
        return redirect(url_for('gestion_trabajadores'))
        
    except Exception as e:
        db.session.rollback()
        flash('❌ Error al actualizar el trabajador: ' + str(e), 'error')
        return redirect(url_for('gestion_trabajadores'))

@app.route('/desactivar_trabajador/<int:trabajador_id>', methods=['POST'])
def desactivar_trabajador(trabajador_id):
    """Desactivar un trabajador"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        trabajador = Trabajador.query.get_or_404(trabajador_id)
        trabajador.activo = False
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/activar_trabajador/<int:trabajador_id>', methods=['POST'])
def activar_trabajador(trabajador_id):
    """Activar un trabajador"""
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        trabajador = Trabajador.query.get_or_404(trabajador_id)
        trabajador.activo = True
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# Agregar esta ruta en app.py
@app.route('/ayuda')
def ayuda():
    return render_template('ayuda.html')

def crear_usuarios_prueba():
    """Crear usuarios de prueba si no existen"""
    with app.app_context():
        if Usuario.query.count() == 0:
            usuarios = [
                {'username': 'admin', 'password': 'admin123', 'rol': 'admin', 'nombre': 'Administrador Principal'},
                {'username': 'compras', 'password': 'compras123', 'rol': 'compras', 'nombre': 'Responsable Compras'},
                {'username': 'stock', 'password': 'stock123', 'rol': 'stock', 'nombre': 'Responsable Stock'},
                {'username': 'basico', 'password': 'basico123', 'rol': 'basico', 'nombre': 'Usuario Básico'}
            ]
            
            for user_data in usuarios:
                usuario = Usuario(
                    username=user_data['username'],
                    rol=user_data['rol'],
                    nombre=user_data['nombre']
                )
                usuario.set_password(user_data['password'])
                db.session.add(usuario)
            
            db.session.commit()
            print("Usuarios de prueba creados exitosamente!")

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        crear_usuarios_prueba()  # ← Esta línea crea los usuarios automáticamente
    app.run(debug=True)